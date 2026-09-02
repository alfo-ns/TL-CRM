"""Export/import of Companies, Contacts and Bridge contacts as an .xlsx
workbook with one sheet per entity. Rows carry their own ID (column A) so a
re-imported file updates existing records instead of duplicating them.

Import is a two-step flow: `preview_import` validates and diffs the workbook
against the current database without writing anything, returning counts and
per-row problems; `apply_import` (called only after the user confirms) runs
the same parsing again and writes inside a single transaction via
db.run_write.
"""
from datetime import date

from openpyxl import Workbook, load_workbook

from constants import ACTION_LABEL, STAGE_INDEX, STAGE_LABEL, STAGE_ORDER, STAGES, ACTION_TYPES

VALID_STAGES = set(STAGE_INDEX)
VALID_ACTIONS = set(ACTION_LABEL)

COMPANY_HEADERS = [
    "ID", "Nome", "Settore", "Dipendenti", "Valore", "Stadio",
    "Portata da", "Gestito da",
    "Prossima azione", "Data prossima azione", "Stadio obiettivo",
    "Bridge collegato",
]
CONTACT_HEADERS = [
    "ID", "ID Azienda", "Nome", "Cognome", "Ruolo", "Email", "Telefono",
    "LinkedIn", "Social", "Social label", "Note",
    "Portato da", "Gestito da",
    "Prossima azione", "Data prossima azione", "Stadio obiettivo",
    "Fonte", "Temperatura", "Potere decisionale", "Orario/canale", "Interessi",
    "Competitor", "Argomenti utili", "Argomenti da evitare", "Eventi/fiere", "Appunti liberi",
]
BRIDGE_HEADERS = [
    "ID", "Nome", "Ruolo", "Relazione", "Email", "Telefono", "LinkedIn", "Note",
    "Aziende collegate",
]

_BRIDGE_SEP = ";"


# ---------------------------------------------------------------- export

def _autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 48)


def build_workbook(conn):
    wb = Workbook()
    wb.remove(wb.active)

    companies_by_id = {r["id"]: r for r in conn.execute("SELECT * FROM companies ORDER BY id")}
    bridges_by_id = {r["id"]: r for r in conn.execute("SELECT * FROM bridges ORDER BY id")}

    ws = wb.create_sheet("Aziende")
    ws.append(COMPANY_HEADERS)
    for c in companies_by_id.values():
        bridge_name = bridges_by_id[c["bridge_id"]]["nome"] if c["bridge_id"] in bridges_by_id else ""
        ws.append([
            c["id"], c["nome"], c["settore"], c["dip"], c["valore"], c["stage"],
            c["portato_da"], c["gestito_da"],
            c["next_tipo"] or "", c["next_data"] or "", c["next_stadio"] or "",
            bridge_name,
        ])
    _autosize(ws)

    ws = wb.create_sheet("Contatti")
    ws.append(CONTACT_HEADERS)
    for p in conn.execute("SELECT * FROM contacts ORDER BY id"):
        ws.append([
            p["id"], p["company_id"], p["nome"], p["cognome"], p["ruolo"], p["email"], p["tel"],
            p["linkedin"], p["social"], p["social_label"], p["note"],
            p["portato_da"], p["gestito_da"],
            p["next_tipo"] or "", p["next_data"] or "", p["next_stadio"] or "",
            p["dossier_fonte"], p["dossier_temperatura"], p["dossier_potere"], p["dossier_orario"],
            p["dossier_interessi"], p["dossier_competitor"], p["dossier_argomenti_utili"],
            p["dossier_argomenti_evitare"], p["dossier_eventi"], p["dossier_libero"],
        ])
    _autosize(ws)

    ws = wb.create_sheet("Bridge")
    ws.append(BRIDGE_HEADERS)
    for b in bridges_by_id.values():
        linked = [
            companies_by_id[r["id"]]["nome"]
            for r in conn.execute("SELECT id FROM companies WHERE bridge_id = ? ORDER BY id", (b["id"],))
            if r["id"] in companies_by_id
        ]
        ws.append([
            b["id"], b["nome"], b["ruolo"], b["relazione"], b["email"], b["tel"], b["linkedin"], b["note"],
            _BRIDGE_SEP.join(linked),
        ])
    _autosize(ws)

    return wb


# ---------------------------------------------------------------- import parsing

def _cell(row, headers, name):
    return row[headers.index(name)]


def _str(v):
    if v is None:
        return ""
    return str(v).strip()


def _int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _date_str(v):
    """openpyxl hands back a datetime for date-formatted cells, or a plain
    string if the cell was typed as text; normalize both to YYYY-MM-DD."""
    if v is None or v == "":
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    return s


def _read_sheet(wb, name, headers):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    file_headers = [_str(h) for h in rows[0]]
    if file_headers[: len(headers)] != headers:
        raise ValueError(
            f"Il foglio '{name}' non ha le colonne attese. "
            f"Riscarica un export recente per avere l'intestazione corretta."
        )
    out = []
    for raw in rows[1:]:
        if raw is None or all(v is None for v in raw):
            continue
        out.append(list(raw) + [None] * (len(headers) - len(raw)))
    return out


class ImportPlan:
    """Result of parsing+validating a workbook, ready to be applied or just
    summarized for the user's preview screen."""

    def __init__(self):
        self.company_updates = []   # (id, fields dict, bridge_name)
        self.company_creates = []   # (fields dict, bridge_name)
        self.contact_updates = []   # (id, fields dict, company_ref)
        self.contact_creates = []   # (fields dict, company_ref)
        self.bridge_updates = []    # (id, fields dict, linked_company_names)
        self.bridge_creates = []    # (fields dict, linked_company_names)
        self.errors = []            # list of {"sheet", "row", "message"}

    def summary(self):
        return {
            "aziende": {"aggiornate": len(self.company_updates), "nuove": len(self.company_creates)},
            "contatti": {"aggiornate": len(self.contact_updates), "nuove": len(self.contact_creates)},
            "bridge": {"aggiornate": len(self.bridge_updates), "nuove": len(self.bridge_creates)},
            "errori": self.errors,
            "ok": len(self.errors) == 0,
        }


def parse_workbook(file_stream, conn):
    """Parse + validate the uploaded workbook against the current DB state.
    Never writes anything. Raises ValueError for structural problems
    (missing/renamed sheets); row-level problems go into plan.errors and
    that row is skipped."""
    wb = load_workbook(file_stream, data_only=True)
    plan = ImportPlan()

    existing_company_ids = {r["id"] for r in conn.execute("SELECT id FROM companies")}
    existing_contact_ids = {r["id"] for r in conn.execute("SELECT id FROM contacts")}
    existing_bridge_ids = {r["id"] for r in conn.execute("SELECT id FROM bridges")}
    company_name_to_id = {r["nome"]: r["id"] for r in conn.execute("SELECT id, nome FROM companies")}

    # ---- Aziende ----
    for i, row in enumerate(_read_sheet(wb, "Aziende", COMPANY_HEADERS), start=2):
        c = dict(zip(COMPANY_HEADERS, row))
        nome = _str(c["Nome"])
        if not nome:
            plan.errors.append({"sheet": "Aziende", "row": i, "message": "Nome mancante"})
            continue
        stage = _str(c["Stadio"]) or "prospect"
        if stage not in VALID_STAGES:
            plan.errors.append({"sheet": "Aziende", "row": i, "message": f"Stadio '{stage}' non valido"})
            continue
        next_tipo = _str(c["Prossima azione"])
        if next_tipo and next_tipo not in VALID_ACTIONS:
            plan.errors.append({"sheet": "Aziende", "row": i, "message": f"Tipo azione '{next_tipo}' non valido"})
            continue
        next_stadio = _str(c["Stadio obiettivo"])
        if next_stadio and next_stadio not in VALID_STAGES:
            plan.errors.append({"sheet": "Aziende", "row": i, "message": f"Stadio obiettivo '{next_stadio}' non valido"})
            continue
        fields = {
            "nome": nome, "settore": _str(c["Settore"]), "dip": _int(c["Dipendenti"]),
            "valore": _int(c["Valore"]), "stage": stage,
            "portato_da": _str(c["Portata da"]), "gestito_da": _str(c["Gestito da"]),
            "next_tipo": next_tipo or None,
            "next_data": _date_str(c["Data prossima azione"]) or None,
            "next_stadio": next_stadio or None,
        }
        row_id = _int(c["ID"], default=None) if c["ID"] not in (None, "") else None
        bridge_name = _str(c["Bridge collegato"])
        if row_id is not None and row_id in existing_company_ids:
            plan.company_updates.append((row_id, fields, bridge_name))
        else:
            plan.company_creates.append((fields, bridge_name))

    # ---- Contatti ----
    for i, row in enumerate(_read_sheet(wb, "Contatti", CONTACT_HEADERS), start=2):
        p = dict(zip(CONTACT_HEADERS, row))
        nome = _str(p["Nome"])
        company_ref = p["ID Azienda"]
        if company_ref in (None, ""):
            plan.errors.append({"sheet": "Contatti", "row": i, "message": "ID Azienda mancante"})
            continue
        company_ref = _int(company_ref, default=None)
        if company_ref is None or company_ref not in existing_company_ids:
            plan.errors.append({"sheet": "Contatti", "row": i, "message": f"Azienda con ID {p['ID Azienda']} non trovata"})
            continue
        next_tipo = _str(p["Prossima azione"])
        if next_tipo and next_tipo not in VALID_ACTIONS:
            plan.errors.append({"sheet": "Contatti", "row": i, "message": f"Tipo azione '{next_tipo}' non valido"})
            continue
        next_stadio = _str(p["Stadio obiettivo"])
        if next_stadio and next_stadio not in VALID_STAGES:
            plan.errors.append({"sheet": "Contatti", "row": i, "message": f"Stadio obiettivo '{next_stadio}' non valido"})
            continue
        fields = {
            "nome": nome or "Nome", "cognome": _str(p["Cognome"]) or "Cognome", "ruolo": _str(p["Ruolo"]),
            "email": _str(p["Email"]), "tel": _str(p["Telefono"]), "linkedin": _str(p["LinkedIn"]),
            "social": _str(p["Social"]), "social_label": _str(p["Social label"]) or "Instagram",
            "note": _str(p["Note"]),
            "portato_da": _str(p["Portato da"]), "gestito_da": _str(p["Gestito da"]),
            "next_tipo": next_tipo or None, "next_data": _date_str(p["Data prossima azione"]) or None,
            "next_stadio": next_stadio or None,
            "dossier_fonte": _str(p["Fonte"]), "dossier_temperatura": _str(p["Temperatura"]) or "Freddo",
            "dossier_potere": _str(p["Potere decisionale"]) or "Da capire", "dossier_orario": _str(p["Orario/canale"]),
            "dossier_interessi": _str(p["Interessi"]), "dossier_competitor": _str(p["Competitor"]),
            "dossier_argomenti_utili": _str(p["Argomenti utili"]),
            "dossier_argomenti_evitare": _str(p["Argomenti da evitare"]),
            "dossier_eventi": _str(p["Eventi/fiere"]), "dossier_libero": _str(p["Appunti liberi"]),
        }
        row_id = _int(p["ID"], default=None) if p["ID"] not in (None, "") else None
        if row_id is not None and row_id in existing_contact_ids:
            plan.contact_updates.append((row_id, fields, company_ref))
        else:
            plan.contact_creates.append((fields, company_ref))

    # ---- Bridge ----
    for i, row in enumerate(_read_sheet(wb, "Bridge", BRIDGE_HEADERS), start=2):
        b = dict(zip(BRIDGE_HEADERS, row))
        nome = _str(b["Nome"])
        if not nome:
            plan.errors.append({"sheet": "Bridge", "row": i, "message": "Nome mancante"})
            continue
        linked_raw = _str(b["Aziende collegate"])
        linked_names = [n.strip() for n in linked_raw.split(_BRIDGE_SEP) if n.strip()]
        unknown = [n for n in linked_names if n not in company_name_to_id]
        if unknown:
            plan.errors.append({
                "sheet": "Bridge", "row": i,
                "message": f"Azienda/e non trovate: {', '.join(unknown)}",
            })
            continue
        fields = {
            "nome": nome, "ruolo": _str(b["Ruolo"]), "relazione": _str(b["Relazione"]) or "Referral",
            "email": _str(b["Email"]), "tel": _str(b["Telefono"]), "linkedin": _str(b["LinkedIn"]),
            "note": _str(b["Note"]),
        }
        row_id = _int(b["ID"], default=None) if b["ID"] not in (None, "") else None
        if row_id is not None and row_id in existing_bridge_ids:
            plan.bridge_updates.append((row_id, fields, linked_names))
        else:
            plan.bridge_creates.append((fields, linked_names))

    return plan


# ---------------------------------------------------------------- import apply

def apply_import(conn, plan, today_iso, today_it):
    """Write plan.*_updates/*_creates to the database. Must run inside the
    caller's db.run_write transaction. Bridge links on companies are resolved
    by name after bridges are created/updated, since a create may need a
    fresh bridge_id that doesn't exist yet."""
    company_id_remap = {}  # index in company_creates -> new id, for reference if ever needed

    for row_id, fields, _bridge_name in plan.company_updates:
        conn.execute(
            "UPDATE companies SET nome=?, settore=?, dip=?, valore=?, stage=?, portato_da=?, gestito_da=?, "
            "next_tipo=?, next_data=?, next_stadio=? WHERE id=?",
            (fields["nome"], fields["settore"], fields["dip"], fields["valore"], fields["stage"],
             fields["portato_da"], fields["gestito_da"], fields["next_tipo"], fields["next_data"],
             fields["next_stadio"], row_id),
        )

    new_company_ids = []
    for fields, _bridge_name in plan.company_creates:
        max_idx = 0 if fields["stage"] == "perso" else STAGE_INDEX[fields["stage"]]
        cur = conn.execute(
            "INSERT INTO companies (nome, settore, dip, valore, stage, max_stage_index, portato_da, gestito_da, "
            "bridge_id, next_tipo, next_data, next_stadio, stage_entered_at, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?, ?, ?)",
            (fields["nome"], fields["settore"], fields["dip"], fields["valore"], fields["stage"], max_idx,
             fields["portato_da"], fields["gestito_da"], fields["next_tipo"], fields["next_data"],
             fields["next_stadio"], today_iso, today_iso),
        )
        new_id = cur.lastrowid
        new_company_ids.append(new_id)
        conn.execute(
            "INSERT INTO stage_history (company_id, stage, entered_at, left_at) VALUES (?, ?, ?, NULL)",
            (new_id, fields["stage"], today_iso),
        )
        conn.execute(
            "INSERT INTO activities (company_id, data_label, tipo, testo, created_at) VALUES (?, ?, ?, ?, ?)",
            (new_id, today_it, "Import", "Azienda creata da importazione XLSX.", today_iso),
        )

    for row_id, fields, _company_ref in plan.contact_updates:
        conn.execute(
            "UPDATE contacts SET nome=?, cognome=?, ruolo=?, email=?, tel=?, linkedin=?, social=?, social_label=?, "
            "note=?, portato_da=?, gestito_da=?, next_tipo=?, next_data=?, next_stadio=?, "
            "dossier_fonte=?, dossier_temperatura=?, dossier_potere=?, dossier_orario=?, dossier_interessi=?, "
            "dossier_competitor=?, dossier_argomenti_utili=?, dossier_argomenti_evitare=?, dossier_eventi=?, "
            "dossier_libero=? WHERE id=?",
            (fields["nome"], fields["cognome"], fields["ruolo"], fields["email"], fields["tel"], fields["linkedin"],
             fields["social"], fields["social_label"], fields["note"], fields["portato_da"], fields["gestito_da"],
             fields["next_tipo"], fields["next_data"], fields["next_stadio"],
             fields["dossier_fonte"], fields["dossier_temperatura"], fields["dossier_potere"], fields["dossier_orario"],
             fields["dossier_interessi"], fields["dossier_competitor"], fields["dossier_argomenti_utili"],
             fields["dossier_argomenti_evitare"], fields["dossier_eventi"], fields["dossier_libero"], row_id),
        )

    for fields, company_ref in plan.contact_creates:
        conn.execute(
            "INSERT INTO contacts (company_id, nome, cognome, ruolo, email, tel, linkedin, social, social_label, "
            "note, created_at, portato_da, gestito_da, next_tipo, next_data, next_stadio, "
            "dossier_fonte, dossier_temperatura, dossier_potere, dossier_orario, dossier_interessi, "
            "dossier_competitor, dossier_argomenti_utili, dossier_argomenti_evitare, dossier_eventi, dossier_libero) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (company_ref, fields["nome"], fields["cognome"], fields["ruolo"], fields["email"], fields["tel"],
             fields["linkedin"], fields["social"], fields["social_label"], fields["note"], today_iso,
             fields["portato_da"], fields["gestito_da"], fields["next_tipo"], fields["next_data"], fields["next_stadio"],
             fields["dossier_fonte"], fields["dossier_temperatura"], fields["dossier_potere"], fields["dossier_orario"],
             fields["dossier_interessi"], fields["dossier_competitor"], fields["dossier_argomenti_utili"],
             fields["dossier_argomenti_evitare"], fields["dossier_eventi"], fields["dossier_libero"]),
        )

    for row_id, fields, _linked_names in plan.bridge_updates:
        conn.execute(
            "UPDATE bridges SET nome=?, ruolo=?, relazione=?, email=?, tel=?, linkedin=?, note=? WHERE id=?",
            (fields["nome"], fields["ruolo"], fields["relazione"], fields["email"], fields["tel"],
             fields["linkedin"], fields["note"], row_id),
        )

    new_bridge_ids = []
    for fields, _linked_names in plan.bridge_creates:
        cur = conn.execute(
            "INSERT INTO bridges (nome, ruolo, relazione, email, tel, linkedin, note, introduzioni) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 0)",
            (fields["nome"], fields["ruolo"], fields["relazione"], fields["email"], fields["tel"],
             fields["linkedin"], fields["note"]),
        )
        new_bridge_ids.append(cur.lastrowid)

    # Resolve bridge <-> company links by name, now that everything has an id.
    company_name_to_id = {r["nome"]: r["id"] for r in conn.execute("SELECT id, nome FROM companies")}
    all_bridges = list(plan.bridge_updates) + list(zip(new_bridge_ids, (f for f, _ in plan.bridge_creates), (n for _, n in plan.bridge_creates)))
    for row_id, _fields, linked_names in all_bridges:
        conn.execute("UPDATE companies SET bridge_id = NULL WHERE bridge_id = ?", (row_id,))
        for name in linked_names:
            cid = company_name_to_id.get(name)
            if cid is not None:
                conn.execute("UPDATE companies SET bridge_id = ? WHERE id = ?", (row_id, cid))
